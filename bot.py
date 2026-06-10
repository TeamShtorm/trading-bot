# ==================================================
# БЛОК 1: ИМПОРТЫ И КОНФИГ
# ==================================================
import asyncio
import sqlite3
import os
from datetime import datetime, timedelta

from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, CallbackQuery, FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton, BotCommand
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from aiohttp import web

BOT_TOKEN = "8803530037:AAHVuMAb6gIzGXBKH8qbteZtFyttz6_hzh0"
DB_NAME = "trades.db"
BT_DB_NAME = "backtests.db"
TRADES_PER_PAGE = 5

# ==================================================
# БЛОК 2: БАЗЫ ДАННЫХ РЕАЛЬНОЙ ТОРГОВЛИ
# ==================================================
def init_dbs():
    conn = sqlite3.connect(DB_NAME)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS trades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            asset TEXT,
            direction TEXT,
            entry_price REAL,
            exit_price REAL,
            volume REAL,
            pnl REAL,
            result TEXT,
            comment TEXT,
            trade_date TEXT,
            links TEXT,
            emotion TEXT
        )
    """)
    conn.execute("CREATE TABLE IF NOT EXISTS users (user_id INTEGER PRIMARY KEY, lang TEXT DEFAULT 'en')")
    conn.commit()
    conn.close()
    init_backtest_db()

def get_user_lang(user_id):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT lang FROM users WHERE user_id = ?", (user_id,))
    r = cur.fetchone()
    conn.close()
    if r and r[0] in ['ru', 'en']:
        return r[0]
    return "en"

def set_user_lang(user_id, lang):
    conn = sqlite3.connect(DB_NAME)
    conn.execute("INSERT OR REPLACE INTO users (user_id, lang) VALUES (?, ?)", (user_id, lang))
    conn.commit()
    conn.close()

def save_trade(user_id, asset, direction, entry_price, exit_price, volume, pnl, result, comment, trade_date, links, emotion):
    conn = sqlite3.connect(DB_NAME)
    conn.execute("""
        INSERT INTO trades (user_id, asset, direction, entry_price, exit_price, volume, pnl, result, comment, trade_date, links, emotion)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (user_id, asset, direction, entry_price, exit_price, volume, pnl, result, comment, trade_date, links, emotion))
    conn.commit()
    conn.close()

def get_trades_filtered(user_id, result_filter=None, asset_filter=None, date_filter=None):
    conn = sqlite3.connect(DB_NAME)
    query = "SELECT * FROM trades WHERE user_id = ?"
    params = [user_id]
    if result_filter and result_filter != "all":
        if result_filter == "take":
            query += " AND result = 'TAKE'"
        elif result_filter == "stop":
            query += " AND result = 'STOP'"
        elif result_filter == "bu":
            query += " AND result = 'BU'"
    if asset_filter:
        query += " AND asset = ?"
        params.append(asset_filter)
    if date_filter:
        days = {"day": 1, "week": 7, "month": 30}
        start = (datetime.now() - timedelta(days=days[date_filter])).strftime("%Y-%m-%d")
        query += " AND trade_date >= ?"
        params.append(start)
    query += " ORDER BY trade_date DESC"
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df

def get_all_assets(user_id):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT asset FROM trades WHERE user_id = ?", (user_id,))
    assets = [row[0] for row in cur.fetchall()]
    conn.close()
    return assets

def delete_trade(trade_id, user_id):
    conn = sqlite3.connect(DB_NAME)
    conn.execute("DELETE FROM trades WHERE id = ? AND user_id = ?", (trade_id, user_id))
    conn.commit()
    conn.close()

def get_trade_by_id(trade_id, user_id):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT * FROM trades WHERE id = ? AND user_id = ?", (trade_id, user_id))
    row = cur.fetchone()
    conn.close()
    if row:
        cols = ['id', 'user_id', 'asset', 'direction', 'entry_price', 'exit_price', 'volume', 'pnl', 'result', 'comment', 'trade_date', 'links', 'emotion']
        return dict(zip(cols, row))
    return None

def clear_trades(user_id):
    conn = sqlite3.connect(DB_NAME)
    conn.execute("DELETE FROM trades WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()

# ==================================================
# БЛОК 3: БАЗЫ ДАННЫХ БЭКТЕСТА
# ==================================================
def init_backtest_db():
    conn = sqlite3.connect(BT_DB_NAME)
    # Таблица периодов бэктеста (с датами начала и конца)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS backtest_periods (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            name TEXT,
            asset TEXT,
            initial_balance REAL,
            period_start TEXT,
            period_end TEXT,
            created_at TEXT
        )
    """)
    # Таблица таймфреймов и ссылок для периода
    conn.execute("""
        CREATE TABLE IF NOT EXISTS backtest_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            period_id INTEGER,
            timeframe TEXT,
            link TEXT,
            FOREIGN KEY (period_id) REFERENCES backtest_periods(id)
        )
    """)
    # Таблица сделок внутри бэктеста
    conn.execute("""
        CREATE TABLE IF NOT EXISTS backtest_trades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            period_id INTEGER,
            trade_date TEXT,
            direction TEXT,
            entry_price REAL,
            exit_price REAL,
            volume REAL,
            pnl REAL,
            result TEXT,
            comment TEXT,
            FOREIGN KEY (period_id) REFERENCES backtest_periods(id)
        )
    """)
    conn.commit()
    conn.close()

def save_backtest_period(user_id, name, asset, initial_balance, period_start, period_end):
    conn = sqlite3.connect(BT_DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO backtest_periods (user_id, name, asset, initial_balance, period_start, period_end, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (user_id, name, asset, initial_balance, period_start, period_end, datetime.now().isoformat()))
    conn.commit()
    period_id = cur.lastrowid
    conn.close()
    return period_id

def save_backtest_link(period_id, timeframe, link):
    conn = sqlite3.connect(BT_DB_NAME)
    conn.execute("""
        INSERT INTO backtest_links (period_id, timeframe, link)
        VALUES (?, ?, ?)
    """, (period_id, timeframe, link))
    conn.commit()
    conn.close()

def get_backtest_links(period_id):
    conn = sqlite3.connect(BT_DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT timeframe, link FROM backtest_links WHERE period_id = ?", (period_id,))
    rows = cur.fetchall()
    conn.close()
    return rows

def get_backtest_periods(user_id):
    conn = sqlite3.connect(BT_DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT * FROM backtest_periods WHERE user_id = ? ORDER BY created_at DESC", (user_id,))
    rows = cur.fetchall()
    conn.close()
    return rows

def get_backtest_period_by_id(period_id, user_id):
    conn = sqlite3.connect(BT_DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT * FROM backtest_periods WHERE id = ? AND user_id = ?", (period_id, user_id))
    row = cur.fetchone()
    conn.close()
    return row

def delete_backtest_period(period_id, user_id):
    conn = sqlite3.connect(BT_DB_NAME)
    conn.execute("DELETE FROM backtest_links WHERE period_id = ?", (period_id,))
    conn.execute("DELETE FROM backtest_trades WHERE period_id = ?", (period_id,))
    conn.execute("DELETE FROM backtest_periods WHERE id = ? AND user_id = ?", (period_id, user_id))
    conn.commit()
    conn.close()

def clear_backtest_periods(user_id):
    conn = sqlite3.connect(BT_DB_NAME)
    # Сначала получаем все period_id пользователя
    cur = conn.cursor()
    cur.execute("SELECT id FROM backtest_periods WHERE user_id = ?", (user_id,))
    period_ids = [row[0] for row in cur.fetchall()]
    
    # Удаляем ссылки и сделки для этих периодов
    for pid in period_ids:
        conn.execute("DELETE FROM backtest_links WHERE period_id = ?", (pid,))
        conn.execute("DELETE FROM backtest_trades WHERE period_id = ?", (pid,))
    
    conn.execute("DELETE FROM backtest_periods WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()

def save_backtest_trade(period_id, trade_date, direction, entry_price, exit_price, volume, pnl, result, comment):
    conn = sqlite3.connect(BT_DB_NAME)
    conn.execute("""
        INSERT INTO backtest_trades (period_id, trade_date, direction, entry_price, exit_price, volume, pnl, result, comment)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (period_id, trade_date, direction, entry_price, exit_price, volume, pnl, result, comment))
    conn.commit()
    conn.close()

def get_backtest_trades(period_id):
    conn = sqlite3.connect(BT_DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT * FROM backtest_trades WHERE period_id = ? ORDER BY trade_date ASC", (period_id,))
    rows = cur.fetchall()
    conn.close()
    return rows

def delete_backtest_trade(trade_id, period_id, user_id):
    conn = sqlite3.connect(BT_DB_NAME)
    conn.execute("DELETE FROM backtest_trades WHERE id = ? AND period_id IN (SELECT id FROM backtest_periods WHERE user_id = ?)", (trade_id, user_id))
    conn.commit()
    conn.close()

def get_backtests(user_id):
    """Получить все бэктесты пользователя в виде DataFrame"""
    periods = get_backtest_periods(user_id)
    if not periods:
        return pd.DataFrame()
    
    data = []
    for period in periods:
        period_id = period[0]
        trades = get_backtest_trades(period_id)
        for trade in trades:
            data.append({
                'period_start': period[5],
                'period_end': period[6],
                'timeframe': 'N/A',
                'asset': period[3],
                'direction': trade[2],
                'entry_price': trade[3],
                'exit_price': trade[4],
                'link_chart': 'N/A'
            })
    
    if not data:
        return pd.DataFrame()
    
    df = pd.DataFrame(data)
    return df

def export_backtest_to_excel(df, user_id):
    """Экспорт бэктестов в Excel"""
    if df.empty:
        return None
    df_exp = df.copy()
    df_exp = df_exp[['period_start', 'period_end', 'timeframe', 'asset', 'direction', 'entry_price', 'exit_price', 'link_chart']]
    df_exp.columns = ['📅 Начало', '📅 Конец', '⏱ Таймфрейм', '🪙 Актив', '📈 Направление', '💰 Вход', '💰 Выход', '🔗 Ссылка']
    df_exp['📈 Направление'] = df_exp['📈 Направление'].replace({'LONG': '🟢 LONG', 'SHORT': '🔴 SHORT'})
    df_exp = df_exp.sort_values('📅 Начало', ascending=False)
    fname = f"backtest_{user_id}.xlsx"
    with pd.ExcelWriter(fname, engine='openpyxl') as w:
        df_exp.to_excel(w, sheet_name='Бэктест', index=False)
        ws = w.sheets['Бэктест']
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2b6cb0", end_color="2b6cb0", fill_type="solid")
        for col in range(1, len(df_exp.columns)+1):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).fill = header_fill
        for col in range(1, len(df_exp.columns)+1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in range(1, len(df_exp)+2):
                v = ws.cell(row=row, column=col).value
                if v:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len+2, 30)
        ws.freeze_panes = 'A2'
    return fname

# ==================================================
# БЛОК 4: ГРАФИК И СТАТИСТИКА
# ==================================================
def generate_equity_chart(df, user_id):
    if df.empty:
        return None
    df = df.sort_values('trade_date')
    cum_pnl = df['pnl'].cumsum()
    
    plt.figure(figsize=(10, 5))
    plt.plot(cum_pnl.values, marker='o', color='#2b6cb0', linewidth=2, markersize=4)
    plt.axhline(0, color='red', linestyle='--', linewidth=1)
    plt.fill_between(range(len(cum_pnl)), 0, cum_pnl.values, where=(cum_pnl.values >= 0), color='green', alpha=0.3)
    plt.fill_between(range(len(cum_pnl)), 0, cum_pnl.values, where=(cum_pnl.values < 0), color='red', alpha=0.3)
    plt.title("Кривая доходности", fontsize=14, fontweight='bold')
    plt.xlabel("Номер сделки")
    plt.ylabel("Накопленный P&L ($)")
    plt.grid(True, linestyle=':', alpha=0.6)
    plt.tight_layout()
    
    fname = f"equity_{user_id}.png"
    plt.savefig(fname, bbox_inches='tight', dpi=100)
    plt.close()
    return fname

def generate_backtest_equity_chart(trades, initial_balance, period_id, period_name=""):
    if not trades:
        return None
    
    balance = [initial_balance]
    for trade in trades:
        new_balance = balance[-1] + trade[6]
        balance.append(new_balance)
    
    plt.figure(figsize=(10, 5))
    plt.plot(balance, marker='o', color='#2b6cb0', linewidth=2, markersize=4)
    plt.axhline(initial_balance, color='red', linestyle='--', linewidth=1, label=f'Начальный баланс: ${initial_balance:.0f}')
    plt.fill_between(range(len(balance)), initial_balance, balance, where=(balance >= initial_balance), color='green', alpha=0.3)
    plt.fill_between(range(len(balance)), initial_balance, balance, where=(balance < initial_balance), color='red', alpha=0.3)
    plt.title(f"Кривая доходности: {period_name}", fontsize=12, fontweight='bold')
    plt.xlabel("Номер сделки")
    plt.ylabel("Баланс ($)")
    plt.legend()
    plt.grid(True, linestyle=':', alpha=0.6)
    plt.tight_layout()
    
    fname = f"backtest_equity_{period_id}.png"
    plt.savefig(fname, bbox_inches='tight', dpi=100)
    plt.close()
    return fname

def get_stats_text(df):
    if df.empty:
        return "Нет данных."
    total = len(df)
    wins = len(df[df['pnl'] > 0])
    losses = len(df[df['pnl'] < 0])
    bu = len(df[df['pnl'] == 0])
    wr = wins/total*100 if total else 0
    longs = len(df[df['direction'] == 'LONG'])
    shorts = len(df[df['direction'] == 'SHORT'])
    total_pnl = df['pnl'].sum()
    avg_pnl = df['pnl'].mean()
    best = df['pnl'].max()
    worst = df['pnl'].min()
    sum_profit = df[df['pnl']>0]['pnl'].sum()
    sum_loss = abs(df[df['pnl']<0]['pnl'].sum())
    pf = sum_profit/sum_loss if sum_loss else sum_profit
    emotions = df['emotion'].value_counts().to_dict()
    emotion_text = "\n".join([f"{e}: {c}" for e, c in emotions.items()]) if emotions else "нет данных"
    return (
        f"Ваша статистика\n\n"
        f"Всего сделок: {total}\n"
        f"Тейков: {wins}\n"
        f"Стопов: {losses}\n"
        f"БУ: {bu}\n"
        f"Винрейт: {wr:.1f}%\n"
        f"Лонги: {longs} | Шорты: {shorts}\n"
        f"Суммарный P&L: ${total_pnl:.2f}\n"
        f"Средняя сделка: ${avg_pnl:.2f}\n"
        f"Лучшая: +${best:.2f}\n"
        f"Худшая: ${worst:.2f}\n"
        f"Профит-фактор: {pf:.2f}\n\n"
        f"Эмоции:\n{emotion_text}"
    )

def get_stats_text_short(df, title):
    if df.empty:
        return f"{title}\n\nНет данных."
    total = len(df)
    wins = len(df[df['pnl'] > 0])
    losses = len(df[df['pnl'] < 0])
    wr = wins/total*100 if total else 0
    total_pnl = df['pnl'].sum()
    return f"{title}\n\nСделок: {total}\nТейков: {wins} | Стопов: {losses}\nВинрейт: {wr:.1f}%\nP&L: ${total_pnl:.2f}"

def get_backtest_stats_text(trades, initial_balance, period_name=""):
    if not trades:
        return f"Статистика бэктеста: {period_name}\n\nНет сделок в этом периоде."
    
    total = len(trades)
    wins = len([t for t in trades if t[6] > 0])
    losses = len([t for t in trades if t[6] < 0])
    wr = wins/total*100 if total else 0
    total_pnl = sum(t[6] for t in trades)
    final_balance = initial_balance + total_pnl
    avg_pnl = total_pnl / total if total else 0
    best = max(t[6] for t in trades) if trades else 0
    worst = min(t[6] for t in trades) if trades else 0
    
    balance = initial_balance
    max_balance = initial_balance
    max_drawdown = 0
    for trade in trades:
        balance += trade[6]
        if balance > max_balance:
            max_balance = balance
        drawdown = (max_balance - balance) / max_balance * 100 if max_balance > 0 else 0
        if drawdown > max_drawdown:
            max_drawdown = drawdown
    
    return (
        f"Статистика бэктеста: {period_name}\n\n"
        f"Начальный баланс: ${initial_balance:.2f}\n"
        f"Конечный баланс: ${final_balance:.2f}\n"
        f"Общий P&L: ${total_pnl:+.2f}\n"
        f"Макс. просадка: {max_drawdown:.1f}%\n\n"
        f"Всего сделок: {total}\n"
        f"Тейков: {wins}\n"
        f"Стопов: {losses}\n"
        f"Винрейт: {wr:.1f}%\n"
        f"Средняя сделка: ${avg_pnl:.2f}\n"
        f"Лучшая сделка: +${best:.2f}\n"
        f"Худшая сделка: ${worst:.2f}"
    )

# ==================================================
# БЛОК 5: EXCEL
# ==================================================
def export_real_to_excel(df, user_id):
    if df.empty:
        return None
    df_exp = df.copy()
    df_exp = df_exp[['trade_date', 'asset', 'direction', 'entry_price', 'exit_price', 'volume', 'pnl', 'result', 'comment', 'links', 'emotion']]
    df_exp.columns = ['📅 Дата', '🪙 Актив', '📈 Направление', '💰 Вход', '💰 Выход', '📊 Объём', '💵 P&L', '🎯 Исход', '📝 Комментарий', '🔗 Ссылки', '😊 Эмоции']
    df_exp['📈 Направление'] = df_exp['📈 Направление'].replace({'LONG': '🟢 LONG', 'SHORT': '🔴 SHORT'})
    df_exp['🎯 Исход'] = df_exp['🎯 Исход'].replace({'TAKE': '✅ Тейк', 'STOP': '❌ Стоп', 'BU': '⚖️ БУ'})
    df_exp = df_exp.sort_values('📅 Дата', ascending=False)
    fname = f"real_{user_id}.xlsx"
    with pd.ExcelWriter(fname, engine='openpyxl') as w:
        df_exp.to_excel(w, sheet_name='Реальная торговля', index=False)
        ws = w.sheets['Реальная торговля']
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2b6cb0", end_color="2b6cb0", fill_type="solid")
        for col in range(1, len(df_exp.columns)+1):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).fill = header_fill
        for col in range(1, len(df_exp.columns)+1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in range(1, len(df_exp)+2):
                v = ws.cell(row=row, column=col).value
                if v:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len+2, 30)
        ws.freeze_panes = 'A2'
    return fname

# ==================================================
# БЛОК 6: КЛАВИАТУРЫ
# ==================================================

# ---------- ОСНОВНЫЕ МЕНЮ ----------
def main_menu(lang):
    if lang == "ru":
        text_real = "📊 Реальная торговля"
        text_backtest = "🔄 Бэктест"
        text_settings = "⚙️ Настройки"
    else:
        text_real = "📊 Real Trading"
        text_backtest = "🔄 Backtest"
        text_settings = "⚙️ Settings"
    
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=text_real, callback_data="mode_real")],
        [InlineKeyboardButton(text=text_backtest, callback_data="mode_backtest")],
        [InlineKeyboardButton(text=text_settings, callback_data="settings_menu")]
    ])

def real_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Add Trade", callback_data="real_add_trade")],
        [InlineKeyboardButton(text="📋 Trade List", callback_data="real_list_trades")],
        [InlineKeyboardButton(text="📊 Statistics", callback_data="real_stats_show")],
        [InlineKeyboardButton(text="📎 Excel", callback_data="real_excel")],
        [InlineKeyboardButton(text="🗑 Clear All", callback_data="real_clear")],
        [InlineKeyboardButton(text="🔙 Back", callback_data="back_to_mode_selection")]
    ])

def backtest_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ New Period", callback_data="backtest_add_period")],
        [InlineKeyboardButton(text="📋 Periods", callback_data="backtest_list_periods")],
        [InlineKeyboardButton(text="📊 Statistics", callback_data="backtest_stats_list")],
        [InlineKeyboardButton(text="📎 Excel", callback_data="backtest_excel_list")],
        [InlineKeyboardButton(text="🔙 Back", callback_data="back_to_mode_selection")]
    ])

def backtest_periods_kb(periods, page, total_pages, action):
    buttons = []
    start = (page - 1) * 5
    end = start + 5
    for p in periods[start:end]:
        period_id = p[0]
        period_name = p[2]
        buttons.append([InlineKeyboardButton(text=f"📊 {period_name}", callback_data=f"{action}_period_{period_id}")])
    
    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton(text="⬅️ Назад", callback_data=f"{action}_page_{page-1}"))
    if page < total_pages:
        nav.append(InlineKeyboardButton(text="➡️ Вперед", callback_data=f"{action}_page_{page+1}"))
    if nav:
        buttons.append(nav)
    
    buttons.append([InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_backtest_menu")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def backtest_period_menu_kb(period_id):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Добавить сделку", callback_data=f"bt_add_trade_{period_id}")],
        [InlineKeyboardButton(text="📊 Статистика периода", callback_data=f"bt_stats_{period_id}")],
        [InlineKeyboardButton(text="📈 График", callback_data=f"bt_chart_{period_id}")],
        [InlineKeyboardButton(text="📎 Excel периода", callback_data=f"bt_excel_{period_id}")],
        [InlineKeyboardButton(text="🗑 Очистить период", callback_data=f"bt_clear_period_{period_id}")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="backtest_list_periods")]
    ])

# ---------- КЛАВИАТУРЫ РЕАЛЬНОЙ ТОРГОВЛИ ----------
def real_direction_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🟢 LONG", callback_data="real_dir_LONG"),
         InlineKeyboardButton(text="🔴 SHORT", callback_data="real_dir_SHORT")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_real_trade")]
    ])

def real_result_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Тейк", callback_data="real_res_TAKE"),
         InlineKeyboardButton(text="❌ Стоп", callback_data="real_res_STOP")],
        [InlineKeyboardButton(text="⚖️ БУ", callback_data="real_res_BU")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_real_trade")]
    ])

def real_emotion_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="😊 Спокойствие", callback_data="real_em_calm")],
        [InlineKeyboardButton(text="😨 Страх", callback_data="real_em_fear")],
        [InlineKeyboardButton(text="😈 Жадность", callback_data="real_em_greed")],
        [InlineKeyboardButton(text="🤬 Тильт", callback_data="real_em_tilt")],
        [InlineKeyboardButton(text="😌 Уверенность", callback_data="real_em_confidence")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_real_trade")]
    ])

def real_link_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔗 Добавить ссылку", callback_data="real_add_link")],
        [InlineKeyboardButton(text="✅ Завершить", callback_data="real_link_done")]
    ])

def real_timeframe_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="15м", callback_data="real_tf_15m"),
         InlineKeyboardButton(text="1ч", callback_data="real_tf_1h"),
         InlineKeyboardButton(text="4ч", callback_data="real_tf_4h")],
        [InlineKeyboardButton(text="1д", callback_data="real_tf_1d"),
         InlineKeyboardButton(text="1н", callback_data="real_tf_1w"),
         InlineKeyboardButton(text="1м", callback_data="real_tf_1M")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_real_trade")]
    ])

# ---------- КЛАВИАТУРЫ БЭКТЕСТА ----------
def backtest_direction_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🟢 LONG", callback_data="bt_dir_LONG"),
         InlineKeyboardButton(text="🔴 SHORT", callback_data="bt_dir_SHORT")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_backtest_trade")]
    ])

def backtest_result_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Тейк", callback_data="bt_res_TAKE"),
         InlineKeyboardButton(text="❌ Стоп", callback_data="bt_res_STOP")],
        [InlineKeyboardButton(text="⚖️ БУ", callback_data="bt_res_BU")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_backtest_trade")]
    ])

def backtest_link_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔗 Добавить ссылку", callback_data="bt_add_link")],
        [InlineKeyboardButton(text="✅ Завершить", callback_data="bt_link_done")]
    ])

def backtest_timeframe_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="15м", callback_data="bt_tf_15m"),
         InlineKeyboardButton(text="1ч", callback_data="bt_tf_1h"),
         InlineKeyboardButton(text="4ч", callback_data="bt_tf_4h")],
        [InlineKeyboardButton(text="1д", callback_data="bt_tf_1d"),
         InlineKeyboardButton(text="1н", callback_data="bt_tf_1w"),
         InlineKeyboardButton(text="1м", callback_data="bt_tf_1M")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_backtest_trade")]
    ])

# ---------- ОБЩИЕ КЛАВИАТУРЫ ----------
def settings_menu_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🌐 Change Language", callback_data="change_lang")],
        [InlineKeyboardButton(text="📞 Support", url="https://t.me/TJsupport_bot")],
        [InlineKeyboardButton(text="🔙 Back", callback_data="back_to_mode_selection")]
    ])

def lang_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🇷🇺 Русский", callback_data="lang_ru")],
        [InlineKeyboardButton(text="🇬🇧 English", callback_data="lang_en")],
        [InlineKeyboardButton(text="🔙 Back", callback_data="settings_menu")]
    ])

def back_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_previous")]
    ])

def confirm_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⚠️ ДА, УДАЛИТЬ", callback_data="clear_yes")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_previous")]
    ])

def cancel_real_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_real_trade")]
    ])

def cancel_backtest_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_backtest_trade")]
    ])

def real_trades_list_kb(trades, page, total_pages):
    buttons = []
    for _, row in trades.iterrows():
        pnl = row['pnl']
        emoji = "✅" if pnl > 0 else ("❌" if pnl < 0 else "⚖️")
        buttons.append([InlineKeyboardButton(text=f"{row['asset']} {emoji} ${pnl:.0f}", callback_data=f"real_view_{row['id']}")])
    
    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton(text="⬅️ Назад", callback_data=f"real_page_{page-1}"))
    if page < total_pages:
        nav.append(InlineKeyboardButton(text="➡️ Вперед", callback_data=f"real_page_{page+1}"))
    if nav:
        buttons.append(nav)
    
    buttons.append([InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_real_menu")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def real_view_trade_kb(trade_id):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🗑 Удалить", callback_data=f"real_del_{trade_id}")],
        [InlineKeyboardButton(text="🔙 К списку", callback_data="real_list_trades")]
    ])

def real_stats_main_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📈 Вся статистика", callback_data="real_stats_all")],
        [InlineKeyboardButton(text="💰 По активам", callback_data="real_stats_by_asset")],
        [InlineKeyboardButton(text="📅 По дате", callback_data="real_stats_by_date")],
        [InlineKeyboardButton(text="😊 По эмоциям", callback_data="real_stats_by_emotion")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_real_menu")]
    ])

def real_stats_assets_kb(assets):
    buttons = [[InlineKeyboardButton(text=a, callback_data=f"real_stats_asset_{a}")] for a in assets]
    buttons.append([InlineKeyboardButton(text="🔙 Назад", callback_data="real_stats_show")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def real_stats_date_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📆 День", callback_data="real_stats_date_day")],
        [InlineKeyboardButton(text="📅 Неделя", callback_data="real_stats_date_week")],
        [InlineKeyboardButton(text="📊 Месяц", callback_data="real_stats_date_month")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="real_stats_show")]
    ])

def real_stats_emotions_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="😊 Спокойствие", callback_data="real_stats_em_calm")],
        [InlineKeyboardButton(text="😨 Страх", callback_data="real_stats_em_fear")],
        [InlineKeyboardButton(text="😈 Жадность", callback_data="real_stats_em_greed")],
        [InlineKeyboardButton(text="🤬 Тильт", callback_data="real_stats_em_tilt")],
        [InlineKeyboardButton(text="😌 Уверенность", callback_data="real_stats_em_confidence")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="real_stats_show")]
    ])

def real_filter_menu_kb(result_filter, asset_filter, date_filter, has_assets):
    buttons = []
    buttons.append([InlineKeyboardButton(text=f"{'✅ ' if result_filter == 'all' else ''}Все", callback_data="real_filter_all")])
    buttons.append([InlineKeyboardButton(text=f"{'✅ ' if result_filter == 'take' else ''}✅ Тейк", callback_data="real_filter_take")])
    buttons.append([InlineKeyboardButton(text=f"{'✅ ' if result_filter == 'stop' else ''}❌ Стоп", callback_data="real_filter_stop")])
    buttons.append([InlineKeyboardButton(text=f"{'✅ ' if result_filter == 'bu' else ''}⚖️ БУ", callback_data="real_filter_bu")])
    if has_assets:
        buttons.append([InlineKeyboardButton(text="💰 По активу", callback_data="real_filter_asset_menu")])
    buttons.append([InlineKeyboardButton(text="📅 По дате", callback_data="real_filter_date_menu")])
    buttons.append([InlineKeyboardButton(text="🗑 Сбросить фильтры", callback_data="real_filter_clear")])
    buttons.append([InlineKeyboardButton(text="🔙 Назад", callback_data="real_list_trades")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def real_filter_asset_kb(assets):
    buttons = [[InlineKeyboardButton(text=a, callback_data=f"real_filter_asset_{a}")] for a in assets]
    buttons.append([InlineKeyboardButton(text="🔙 Назад", callback_data="real_filter_menu")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def real_filter_date_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📆 День", callback_data="real_filter_date_day")],
        [InlineKeyboardButton(text="📅 Неделя", callback_data="real_filter_date_week")],
        [InlineKeyboardButton(text="📊 Месяц", callback_data="real_filter_date_month")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="real_filter_menu")]
    ])

# ==================================================
# БЛОК 7: FSM
# ==================================================
class TradeForm(StatesGroup):
    asset = State()
    direction = State()
    entry_price = State()
    exit_price = State()
    volume = State()
    result = State()
    comment = State()
    add_link = State()
    link_tf = State()
    link_url = State()
    links = State()
    trade_date = State()
    emotion = State()

class BacktestForm(StatesGroup):
    period_start = State()
    period_end = State()
    timeframe = State()
    asset = State()
    direction = State()
    entry_price = State()
    exit_price = State()
    link_chart = State()

class BacktestPeriodForm(StatesGroup):
    name = State()
    asset = State()
    period_start = State()
    period_end = State()
    initial_balance = State()

class BacktestTradeForm(StatesGroup):
    period_id = State()
    trade_date = State()
    direction = State()
    entry_price = State()
    exit_price = State()
    volume = State()
    result = State()
    comment = State()
    add_link = State()
    link_timeframe = State()
    link_url = State()

# ==================================================
# БЛОК 8: ВЕБ-СЕРВЕР ДЛЯ RENDER
# ==================================================

async def health_check(request):
    return web.Response(text="Bot is alive!")

async def run_web_server():
    app = web.Application()
    app.router.add_get('/', health_check)
    app.router.add_get('/health', health_check)
    runner = web.AppRunner(app)
    await runner.setup()
    port = int(os.environ.get("PORT", 10000))
    site = web.TCPSite(runner, '0.0.0.0', port)
    await site.start()
    print(f"🌐 Веб-сервер запущен на порту {port}")
    await asyncio.Event().wait()

# ==================================================
# БЛОК 9: ОБРАБОТЧИКИ СТАРТ И НАСТРОЙКИ
# ==================================================

def parse_number(text: str) -> float:
    """Очищает строку и преобразует в число"""
    cleaned = text.strip().replace(",", ".").replace(" ", "")
    return float(cleaned)

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

@dp.message(CommandStart())
async def cmd_start(msg: Message, state: FSMContext):
    await state.clear()
    uid = msg.from_user.id
    lang = get_user_lang(uid)
    if lang == "en":
        await msg.answer("🎛 Select mode:", reply_markup=main_menu(lang))
    else:
        await msg.answer("🎛 Выберите режим работы:", reply_markup=main_menu(lang))

@dp.callback_query(F.data.startswith("lang_"))
async def set_lang(call: CallbackQuery, state: FSMContext):
    lang = call.data.split("_")[1]
    set_user_lang(call.from_user.id, lang)
    await state.clear()
    await call.message.delete()
    if lang == "en":
        await call.message.answer("🎛 Select mode:", reply_markup=main_menu(lang))
    else:
        await call.message.answer("🎛 Выберите режим работы:", reply_markup=main_menu(lang))
    await call.answer()

@dp.callback_query(F.data == "mode_real")
async def mode_real(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("📊 Реальная торговля\n\nВыберите действие:", reply_markup=real_menu())
    await call.answer()

@dp.callback_query(F.data == "mode_backtest")
async def mode_backtest(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("🔄 Бэктест\n\nВыберите действие:", reply_markup=backtest_menu())
    await call.answer()

@dp.callback_query(F.data == "back_to_mode_selection")
async def back_to_mode_selection(call: CallbackQuery, state: FSMContext):
    await state.clear()
    lang = get_user_lang(call.from_user.id)
    await call.message.edit_text("🎛 Выберите режим работы:", reply_markup=main_menu(lang))
    await call.answer()

@dp.callback_query(F.data == "back_to_real_menu")
async def back_to_real_menu(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("📊 Реальная торговля\n\nВыберите действие:", reply_markup=real_menu())
    await call.answer()

@dp.callback_query(F.data == "back_to_backtest_menu")
async def back_to_backtest_menu(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("🔄 Бэктест\n\nВыберите действие:", reply_markup=backtest_menu())
    await call.answer()

@dp.callback_query(F.data == "settings_menu")
async def settings_menu(call: CallbackQuery):
    await call.message.edit_text("⚙️ Настройки\n\nВыберите действие:", reply_markup=settings_menu_kb())
    await call.answer()

@dp.callback_query(F.data == "change_lang")
async def change_lang(call: CallbackQuery):
    await call.message.edit_text("🌐 Select language:", reply_markup=lang_kb())
    await call.answer()

# ---------- ОТМЕНА НЕЗАВЕРШЁННОЙ СДЕЛКИ (РЕАЛЬНАЯ ТОРГОВЛЯ) ----------
@dp.callback_query(F.data == "cancel_real_trade")
async def cancel_real_trade(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("❌ Добавление сделки отменено. Возврат в меню реальной торговли.", reply_markup=real_menu())
    await call.answer()

# ---------- ОТМЕНА НЕЗАВЕРШЁННОЙ СДЕЛКИ (БЭКТЕСТ) ----------
@dp.callback_query(F.data == "cancel_backtest_trade")
async def cancel_backtest_trade(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("❌ Добавление сделки отменено. Возврат в меню бэктеста.", reply_markup=backtest_menu())
    await call.answer()

# ==================================================
# БЛОК 10: ОБРАБОТЧИКИ РЕАЛЬНОЙ ТОРГОВЛИ (ДОБАВЛЕНИЕ СДЕЛКИ)
# ==================================================

@dp.callback_query(F.data == "real_add_trade")
async def real_add_trade(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(TradeForm.asset)
    await call.message.edit_text("📝 Введите тикер (BTC, ETH, TON, AAPL):", reply_markup=cancel_real_kb())
    await call.answer()

@dp.message(TradeForm.asset)
async def real_asset(msg: Message, state: FSMContext):
    await state.update_data(asset=msg.text.upper())
    await state.set_state(TradeForm.direction)
    await msg.answer("📈 Выберите направление:", reply_markup=real_direction_kb())

@dp.callback_query(F.data == "real_dir_LONG")
async def real_direction_long(call: CallbackQuery, state: FSMContext):
    await state.update_data(direction="LONG")
    await state.set_state(TradeForm.entry_price)
    await call.message.edit_text("💰 Введите цену входа:", reply_markup=cancel_real_kb())
    await call.answer()

@dp.callback_query(F.data == "real_dir_SHORT")
async def real_direction_short(call: CallbackQuery, state: FSMContext):
    await state.update_data(direction="SHORT")
    await state.set_state(TradeForm.entry_price)
    await call.message.edit_text("💰 Введите цену входа:", reply_markup=cancel_real_kb())
    await call.answer()

@dp.message(TradeForm.entry_price)
async def real_entry(msg: Message, state: FSMContext):
    try:
        val = parse_number(msg.text)
        if val <= 0:
            raise ValueError
        await state.update_data(entry_price=val)
        await state.set_state(TradeForm.exit_price)
        await msg.answer("💰 Введите цену выхода:", reply_markup=cancel_real_kb())
    except ValueError:
        await msg.answer("❌ Ошибка! Введите число (например: 56700.50 или 0.25)", reply_markup=cancel_real_kb())

@dp.message(TradeForm.exit_price)
async def real_exit(msg: Message, state: FSMContext):
    try:
        val = parse_number(msg.text)
        if val <= 0:
            raise ValueError
        await state.update_data(exit_price=val)
        await state.set_state(TradeForm.volume)
        await msg.answer("📊 Введите объём позиции (например: 0.25 или 1000):", reply_markup=cancel_real_kb())
    except ValueError:
        await msg.answer("❌ Ошибка! Введите число (например: 56700.50 или 0.25)", reply_markup=cancel_real_kb())

@dp.message(TradeForm.volume)
async def real_volume(msg: Message, state: FSMContext):
    try:
        vol = parse_number(msg.text)
        if vol <= 0:
            raise ValueError
        data = await state.get_data()
        direction = data['direction']
        entry = data['entry_price']
        exit_p = data['exit_price']
        if direction == "LONG":
            pnl = (exit_p - entry) * vol
        else:
            pnl = (entry - exit_p) * vol
        await state.update_data(volume=vol, pnl=pnl)
        await state.set_state(TradeForm.result)
        await msg.answer("🎯 Как закрылась сделка?", reply_markup=real_result_kb())
    except ValueError:
        await msg.answer("❌ Ошибка! Введите число (например: 0.25 или 1000)", reply_markup=cancel_real_kb())

@dp.callback_query(F.data == "real_res_TAKE")
async def real_result_take(call: CallbackQuery, state: FSMContext):
    await state.update_data(result="TAKE")
    await state.set_state(TradeForm.comment)
    await call.message.edit_text("📝 Введите комментарий (отправьте '-' чтобы пропустить):", reply_markup=cancel_real_kb())
    await call.answer()

@dp.callback_query(F.data == "real_res_STOP")
async def real_result_stop(call: CallbackQuery, state: FSMContext):
    await state.update_data(result="STOP")
    await state.set_state(TradeForm.comment)
    await call.message.edit_text("📝 Введите комментарий (отправьте '-' чтобы пропустить):", reply_markup=cancel_real_kb())
    await call.answer()

@dp.callback_query(F.data == "real_res_BU")
async def real_result_bu(call: CallbackQuery, state: FSMContext):
    await state.update_data(result="BU")
    await state.update_data(pnl=0)
    await state.set_state(TradeForm.comment)
    await call.message.edit_text("📝 Введите комментарий (отправьте '-' чтобы пропустить):", reply_markup=cancel_real_kb())
    await call.answer()

@dp.message(TradeForm.comment)
async def real_comment(msg: Message, state: FSMContext):
    com = msg.text.strip()
    await state.update_data(comment="" if com == "-" else com)
    await state.set_state(TradeForm.add_link)
    await msg.answer("🔗 Хотите добавить ссылку на график?", reply_markup=real_link_kb())

@dp.callback_query(F.data == "real_add_link")
async def real_add_link_yes(call: CallbackQuery, state: FSMContext):
    await state.set_state(TradeForm.link_url)
    await call.message.edit_text("🔗 Отправьте ссылку:", reply_markup=cancel_real_kb())
    await call.answer()

@dp.callback_query(F.data == "real_link_done")
async def real_link_done(call: CallbackQuery, state: FSMContext):
    await state.update_data(links="")
    await state.set_state(TradeForm.trade_date)
    await call.message.edit_text("📅 Введите дату (ДД.ММ.ГГГГ) или 'сегодня':", reply_markup=cancel_real_kb())
    await call.answer()

@dp.message(TradeForm.link_url)
async def real_get_link(msg: Message, state: FSMContext):
    await state.update_data(link_url=msg.text)
    await state.set_state(TradeForm.link_tf)
    await msg.answer("⏱ Какой это таймфрейм? (15м, 1ч, 4ч, 1д, 1н, 1м):", reply_markup=cancel_real_kb())

@dp.message(TradeForm.link_tf)
async def real_get_tf(msg: Message, state: FSMContext):
    tf = msg.text
    data = await state.get_data()
    links = data.get("links", "")
    new_link = f"{tf}: {data.get('link_url')}"
    links = f"{links}\n{new_link}" if links else new_link
    await state.update_data(links=links)
    await state.set_state(TradeForm.add_link)
    await msg.answer("✅ Ссылка сохранена! Хотите добавить ещё?", reply_markup=real_link_kb())

@dp.message(TradeForm.trade_date)
async def real_date(msg: Message, state: FSMContext):
    dstr = msg.text.strip().lower()
    if dstr in ["сегодня", "today"]:
        trade_date = datetime.now().strftime("%Y-%m-%d")
    else:
        try:
            trade_date = datetime.strptime(dstr, "%d.%m.%Y").strftime("%Y-%m-%d")
        except ValueError:
            await msg.answer("❌ Ошибка! Введите дату в формате ДД.ММ.ГГГГ", reply_markup=cancel_real_kb())
            return
    await state.update_data(trade_date=trade_date)
    await state.set_state(TradeForm.emotion)
    await msg.answer("😊 Какие эмоции были?", reply_markup=real_emotion_kb())

@dp.callback_query(F.data == "real_em_calm")
async def real_emotion_calm(call: CallbackQuery, state: FSMContext):
    await finish_real_trade(call, state, "Спокойствие")

@dp.callback_query(F.data == "real_em_fear")
async def real_emotion_fear(call: CallbackQuery, state: FSMContext):
    await finish_real_trade(call, state, "Страх")

@dp.callback_query(F.data == "real_em_greed")
async def real_emotion_greed(call: CallbackQuery, state: FSMContext):
    await finish_real_trade(call, state, "Жадность")

@dp.callback_query(F.data == "real_em_tilt")
async def real_emotion_tilt(call: CallbackQuery, state: FSMContext):
    await finish_real_trade(call, state, "Тильт")

@dp.callback_query(F.data == "real_em_confidence")
async def real_emotion_confidence(call: CallbackQuery, state: FSMContext):
    await finish_real_trade(call, state, "Уверенность")

async def finish_real_trade(call: CallbackQuery, state: FSMContext, emotion: str):
    data = await state.get_data()
    save_trade(
        user_id=call.from_user.id,
        asset=data['asset'],
        direction=data['direction'],
        entry_price=data['entry_price'],
        exit_price=data['exit_price'],
        volume=data['volume'],
        pnl=data['pnl'],
        result=data['result'],
        comment=data['comment'],
        trade_date=data['trade_date'],
        links=data.get('links', ''),
        emotion=emotion
    )
    await state.clear()
    await call.message.edit_text("✅ Сделка сохранена!", reply_markup=real_menu())
    await call.answer()

# ==================================================
# БЛОК 11: ОБРАБОТЧИКИ СПИСКА СДЕЛОК (ФИЛЬТРЫ И ПАГИНАЦИЯ)
# ==================================================

@dp.callback_query(F.data == "real_list_trades")
async def real_list_trades(call: CallbackQuery, state: FSMContext):
    await state.update_data(page=1, result_filter="all", asset_filter=None, date_filter=None)
    await show_trades_page(call, state)

async def show_trades_page(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    page = data.get('page', 1)
    result_filter = data.get('result_filter', 'all')
    asset_filter = data.get('asset_filter', None)
    date_filter = data.get('date_filter', None)
    
    df = get_trades_filtered(call.from_user.id, result_filter, asset_filter, date_filter)
    if df.empty:
        await call.answer("Нет данных", show_alert=True)
        return
    
    total = len(df)
    total_pages = (total + TRADES_PER_PAGE - 1) // TRADES_PER_PAGE
    if page > total_pages:
        page = total_pages
    start = (page - 1) * TRADES_PER_PAGE
    end = start + TRADES_PER_PAGE
    trades_df = df.iloc[start:end]
    
    await state.update_data(page=page)
    
    filter_text = ""
    if result_filter != "all":
        filter_text += f" | Фильтр: {result_filter}"
    if asset_filter:
        filter_text += f" | {asset_filter}"
    if date_filter:
        filter_text += f" | {date_filter}"
    
    text = f"Сделки (страница {page}/{total_pages}){filter_text}"
    await call.message.edit_text(text, reply_markup=real_trades_list_kb(trades_df, page, total_pages))
    await call.answer()

@dp.callback_query(F.data.startswith("real_page_"))
async def real_change_page(call: CallbackQuery, state: FSMContext):
    page = int(call.data.split("_")[2])
    await state.update_data(page=page)
    await show_trades_page(call, state)

@dp.callback_query(F.data == "real_filter_menu")
async def real_filter_menu(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    result_filter = data.get('result_filter', 'all')
    asset_filter = data.get('asset_filter', None)
    date_filter = data.get('date_filter', None)
    assets = get_all_assets(call.from_user.id)
    await call.message.edit_text("Фильтры:", reply_markup=real_filter_menu_kb(result_filter, asset_filter, date_filter, len(assets) > 0))
    await call.answer()

@dp.callback_query(F.data.startswith("real_filter_"))
async def real_apply_filter(call: CallbackQuery, state: FSMContext):
    action = call.data.split("_")[2]
    
    if action == "all":
        await state.update_data(result_filter="all", asset_filter=None, date_filter=None, page=1)
    elif action == "take":
        await state.update_data(result_filter="take", page=1)
    elif action == "stop":
        await state.update_data(result_filter="stop", page=1)
    elif action == "bu":
        await state.update_data(result_filter="bu", page=1)
    elif action == "clear":
        await state.update_data(result_filter="all", asset_filter=None, date_filter=None, page=1)
        await show_trades_page(call, state)
        return
    elif action == "asset":
        await call.message.edit_text("Выберите актив:", reply_markup=real_filter_asset_kb(get_all_assets(call.from_user.id)))
        return
    elif action == "date":
        await call.message.edit_text("Выберите период:", reply_markup=real_filter_date_kb())
        return
    
    await show_trades_page(call, state)

@dp.callback_query(F.data.startswith("real_filter_asset_"))
async def real_apply_asset_filter(call: CallbackQuery, state: FSMContext):
    asset = call.data.split("_")[3]
    await state.update_data(asset_filter=asset, page=1)
    await show_trades_page(call, state)

@dp.callback_query(F.data.startswith("real_filter_date_"))
async def real_apply_date_filter(call: CallbackQuery, state: FSMContext):
    date_filter = call.data.split("_")[3]
    await state.update_data(date_filter=date_filter, page=1)
    await show_trades_page(call, state)

# ========== ПРОСМОТР СДЕЛКИ ==========
@dp.callback_query(F.data.startswith("real_view_"))
async def real_view_trade(call: CallbackQuery):
    trade_id = int(call.data.split("_")[2])
    trade = get_trade_by_id(trade_id, call.from_user.id)
    if not trade:
        await call.answer("Сделка не найдена", show_alert=True)
        return
    links = trade.get('links', '') or '-'
    dir_emoji = "🟢" if trade['direction'] == "LONG" else "🔴"
    result_text = {"TAKE": "Тейк", "STOP": "Стоп", "BU": "БУ"}.get(trade['result'], trade['result'])
    text = (
        f"Сделка #{trade['id']}\n\n"
        f"Актив: {trade['asset']}\n"
        f"Направление: {dir_emoji} {trade['direction']}\n"
        f"Вход: ${trade['entry_price']}\n"
        f"Выход: ${trade['exit_price']}\n"
        f"Объём: {trade['volume']}\n"
        f"P&L: ${trade['pnl']}\n"
        f"Исход: {result_text}\n"
        f"Дата: {trade['trade_date']}\n"
        f"Эмоции: {trade['emotion']}\n"
        f"Ссылки:\n{links}\n"
        f"Комментарий: {trade['comment'] or '-'}"
    )
    await call.message.edit_text(text, reply_markup=real_view_trade_kb(trade_id))
    await call.answer()

# ========== УДАЛЕНИЕ СДЕЛКИ ==========
@dp.callback_query(F.data.startswith("real_del_"))
async def real_delete_confirm(call: CallbackQuery, state: FSMContext):
    trade_id = int(call.data.split("_")[2])
    await state.update_data(delete_id=trade_id)
    await call.message.edit_text(f"⚠️ Удалить сделку #{trade_id}?", reply_markup=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Да", callback_data="real_del_yes"),
         InlineKeyboardButton(text="Нет", callback_data="real_list_trades")]
    ]))
    await call.answer()

@dp.callback_query(F.data == "real_del_yes")
async def real_delete_execute(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    trade_id = data.get('delete_id')
    if trade_id:
        delete_trade(trade_id, call.from_user.id)
    await state.clear()
    await call.message.edit_text("Сделка удалена!", reply_markup=real_menu())
    await call.answer()

# ==================================================
# БЛОК 12: ОБРАБОТЧИКИ СТАТИСТИКИ
# ==================================================

@dp.callback_query(F.data == "real_stats_show")
async def real_stats_menu(call: CallbackQuery):
    await call.message.edit_text("📊 Выберите тип статистики:", reply_markup=real_stats_main_kb())
    await call.answer()

@dp.callback_query(F.data == "real_stats_all")
async def real_stats_all(call: CallbackQuery):
    df = get_trades_filtered(call.from_user.id)
    if df.empty:
        await call.message.edit_text("📭 Нет данных для статистики.", reply_markup=real_stats_main_kb())
        return
    text = get_stats_text(df)
    chart = generate_equity_chart(df, call.from_user.id)
    if chart:
        await call.message.answer_photo(photo=FSInputFile(chart), caption=text)
        os.remove(chart)
    else:
        await call.message.answer(text)
    await call.message.answer("📊 Выберите тип статистики:", reply_markup=real_stats_main_kb())
    await call.answer()

@dp.callback_query(F.data == "real_stats_by_asset")
async def real_stats_by_asset_menu(call: CallbackQuery):
    assets = get_all_assets(call.from_user.id)
    if not assets:
        await call.message.edit_text("📭 Нет активов для статистики.", reply_markup=real_stats_main_kb())
        return
    await call.message.edit_text("💰 Выберите актив:", reply_markup=real_stats_assets_kb(assets))
    await call.answer()

@dp.callback_query(F.data.startswith("real_stats_asset_"))
async def real_stats_asset_show(call: CallbackQuery):
    asset = call.data.split("_")[3]
    df = get_trades_filtered(call.from_user.id, asset_filter=asset)
    if df.empty:
        await call.message.edit_text(f"📭 Нет данных по активу {asset}.", reply_markup=real_stats_assets_kb(get_all_assets(call.from_user.id)))
        return
    text = get_stats_text(df)
    chart = generate_equity_chart(df, call.from_user.id)
    if chart:
        await call.message.answer_photo(photo=FSInputFile(chart), caption=text)
        os.remove(chart)
    else:
        await call.message.answer(text)
    await call.message.answer("💰 Выберите актив:", reply_markup=real_stats_assets_kb(get_all_assets(call.from_user.id)))
    await call.answer()

@dp.callback_query(F.data == "real_stats_by_date")
async def real_stats_by_date_menu(call: CallbackQuery):
    await call.message.edit_text("📅 Выберите период:", reply_markup=real_stats_date_kb())
    await call.answer()

@dp.callback_query(F.data.startswith("real_stats_date_"))
async def real_stats_date_show(call: CallbackQuery):
    period = call.data.split("_")[3]
    df = get_trades_filtered(call.from_user.id, date_filter=period)
    titles = {"day": "📆 Статистика за сегодня", "week": "📅 Статистика за неделю", "month": "📊 Статистика за месяц"}
    if df.empty:
        await call.message.edit_text(f"📭 {titles.get(period, 'Статистика')}\n\nНет данных.", reply_markup=real_stats_date_kb())
        return
    text = get_stats_text_short(df, titles.get(period, "Статистика"))
    await call.message.edit_text(text, reply_markup=real_stats_date_kb())
    await call.answer()

@dp.callback_query(F.data == "real_stats_by_emotion")
async def real_stats_by_emotion_menu(call: CallbackQuery):
    await call.message.edit_text("😊 Выберите эмоцию:", reply_markup=real_stats_emotions_kb())
    await call.answer()

@dp.callback_query(F.data == "real_stats_em_calm")
async def real_stats_emotion_calm(call: CallbackQuery):
    await show_emotion_stats(call, "Спокойствие")

@dp.callback_query(F.data == "real_stats_em_fear")
async def real_stats_emotion_fear(call: CallbackQuery):
    await show_emotion_stats(call, "Страх")

@dp.callback_query(F.data == "real_stats_em_greed")
async def real_stats_emotion_greed(call: CallbackQuery):
    await show_emotion_stats(call, "Жадность")

@dp.callback_query(F.data == "real_stats_em_tilt")
async def real_stats_emotion_tilt(call: CallbackQuery):
    await show_emotion_stats(call, "Тильт")

@dp.callback_query(F.data == "real_stats_em_confidence")
async def real_stats_emotion_confidence(call: CallbackQuery):
    await show_emotion_stats(call, "Уверенность")

async def show_emotion_stats(call: CallbackQuery, emotion: str):
    df = get_trades_filtered(call.from_user.id)
    df = df[df['emotion'] == emotion]
    if df.empty:
        await call.message.edit_text(f"😊 Статистика по эмоции: {emotion}\n\nНет сделок с этой эмоцией.", reply_markup=real_stats_emotions_kb())
        return
    text = get_stats_text(df)
    await call.message.edit_text(text, reply_markup=real_stats_emotions_kb())
    await call.answer()

# ==================================================
# БЛОК 13: ОБРАБОТЧИКИ EXCEL И ОЧИСТКИ
# ==================================================
@dp.callback_query(F.data == "real_excel")
async def real_excel(call: CallbackQuery):
    df = get_trades_filtered(call.from_user.id)
    if df.empty:
        await call.answer("📭 Нет данных", show_alert=True)
        return
    fname = export_real_to_excel(df, call.from_user.id)
    if fname:
        await call.message.answer_document(document=FSInputFile(fname), caption="📊 Ваш отчёт (реальная торговля)")
        os.remove(fname)
    await call.answer()

@dp.callback_query(F.data == "real_clear")
async def real_clear_confirm(call: CallbackQuery):
    await call.message.edit_text("⚠️ Удалить ВСЕ сделки реальной торговли?", reply_markup=confirm_kb())
    await call.answer()

@dp.callback_query(F.data == "clear_yes")
async def clear_yes(call: CallbackQuery):
    clear_trades(call.from_user.id)
    await call.message.edit_text("🗑 Журнал реальной торговли очищен!", reply_markup=real_menu())
    await call.answer()

# ==================================================
# БЛОК 14: ОБРАБОТЧИКИ БЭКТЕСТА
# ==================================================

# ---------- НАВИГАЦИЯ ----------
@dp.callback_query(F.data == "back_to_period_menu")
async def back_to_period_menu(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    period_id = data.get('period_id')
    if period_id:
        await bt_view_period(call, state)
    else:
        await back_to_backtest_menu(call, state)

# ---------- СПИСОК ПЕРИОДОВ ----------
@dp.callback_query(F.data == "backtest_list_periods")
async def bt_list_periods(call: CallbackQuery, state: FSMContext):
    periods = get_backtest_periods(call.from_user.id)
    if not periods:
        await call.answer("📭 Нет периодов", show_alert=True)
        return
    await state.update_data(view_page=1)
    await show_periods_list(call, state, "view")

async def show_periods_list(call: CallbackQuery, state: FSMContext, action: str):
    data = await state.get_data()
    page = data.get(f'{action}_page', 1)
    periods = get_backtest_periods(call.from_user.id)
    if not periods:
        await call.answer("📭 Нет периодов", show_alert=True)
        return
    total_pages = (len(periods) + 4) // 5
    if page > total_pages:
        page = total_pages
    if page < 1:
        page = 1
    await state.update_data({f'{action}_page': page})
    text = f"📋 Периоды бэктестов - страница {page}/{total_pages}:\n\nВыберите период:"
    await call.message.edit_text(text, reply_markup=backtest_periods_kb(periods, page, total_pages, action))
    await call.answer()

@dp.callback_query(F.data.startswith("view_page_"))
async def view_periods_page(call: CallbackQuery, state: FSMContext):
    page = int(call.data.split("_")[2])
    await state.update_data(view_page=page)
    await show_periods_list(call, state, "view")

# ========== ПРОСМОТР ПЕРИОДА С ПАГИНАЦИЕЙ ==========
@dp.callback_query(F.data.startswith("view_period_"))
async def bt_view_period(call: CallbackQuery, state: FSMContext):
    period_id = int(call.data.split("_")[2])
    await state.update_data(period_id=period_id, trade_page=1)
    await show_period_trades(call, state)

async def show_period_trades(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    period_id = data.get('period_id')
    page = data.get('trade_page', 1)
    
    period = get_backtest_period_by_id(period_id, call.from_user.id)
    if not period:
        await call.answer("❌ Период не найден", show_alert=True)
        return
    
    trades = get_backtest_trades(period_id)
    total = len(trades)
    total_pages = (total + 4) // 5 if total > 0 else 1
    
    start = (page - 1) * 5
    end = start + 5
    page_trades = trades[start:end]
    
    text = f"📊 Период: {period[2]}\n"
    text += f"🪙 {period[3]} | 💰 ${period[4]:.2f}\n"
    text += f"📅 {period[5]} — {period[6]}\n\n"
    
    if total > 0:
        total_pnl = sum(t[6] for t in trades)
        final_balance = period[4] + total_pnl
        text += f"📈 Итог: ${final_balance:.2f} | P&L: ${total_pnl:+.2f}\n\n"
    
    text += f"📋 Сделки (стр. {page}/{total_pages}):\n\n"
    
    if page_trades:
        for t in page_trades:
            emoji = "✅" if t[6] > 0 else ("❌" if t[6] < 0 else "⚖️")
            text += f"{emoji} {t[1]} | {t[2]} | ${t[6]:.2f}\n"
    else:
        text += "📭 Нет сделок в этом периоде.\n"
    
    # Кнопки пагинации
    buttons = []
    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton(text="⬅️ Назад", callback_data=f"bt_page_{page-1}"))
    if page < total_pages:
        nav.append(InlineKeyboardButton(text="➡️ Вперед", callback_data=f"bt_page_{page+1}"))
    if nav:
        buttons.append(nav)
    
    buttons.append([InlineKeyboardButton(text="➕ Добавить сделку", callback_data=f"bt_add_trade_{period_id}")])
    buttons.append([InlineKeyboardButton(text="📊 Статистика периода", callback_data=f"bt_stats_{period_id}")])
    buttons.append([InlineKeyboardButton(text="📈 График", callback_data=f"bt_chart_{period_id}")])
    buttons.append([InlineKeyboardButton(text="📎 Excel периода", callback_data=f"bt_excel_{period_id}")])
    buttons.append([InlineKeyboardButton(text="🗑 Очистить период", callback_data=f"bt_clear_period_{period_id}")])
    buttons.append([InlineKeyboardButton(text="🔙 К списку периодов", callback_data="backtest_list_periods")])
    
    await call.message.edit_text(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await call.answer()

@dp.callback_query(F.data.startswith("bt_page_"))
async def bt_change_page(call: CallbackQuery, state: FSMContext):
    page = int(call.data.split("_")[2])
    await state.update_data(trade_page=page)
    await show_period_trades(call, state)

# ========== СТАТИСТИКА ПЕРИОДА ==========
@dp.callback_query(F.data.startswith("bt_stats_"))
async def bt_show_period_stats(call: CallbackQuery, state: FSMContext):
    period_id = int(call.data.split("_")[2])
    period = get_backtest_period_by_id(period_id, call.from_user.id)
    if not period:
        await call.answer("❌ Период не найден", show_alert=True)
        return
    trades = get_backtest_trades(period_id)
    text = get_backtest_stats_text(trades, period[4], period[2])
    await call.message.edit_text(text, reply_markup=backtest_period_menu_kb(period_id))
    await call.answer()

# ========== EXCEL ПЕРИОДА ==========
@dp.callback_query(F.data.startswith("bt_excel_"))
async def bt_export_period_excel(call: CallbackQuery):
    period_id = int(call.data.split("_")[2])
    period = get_backtest_period_by_id(period_id, call.from_user.id)
    if not period:
        await call.answer("❌ Период не найден", show_alert=True)
        return
    trades = get_backtest_trades(period_id)
    if not trades:
        await call.answer("📭 Нет сделок для экспорта", show_alert=True)
        return
    
    data = []
    for t in trades:
        data.append({
            'trade_date': t[1],
            'direction': t[2],
            'entry_price': t[3],
            'exit_price': t[4],
            'volume': t[5],
            'pnl': t[6],
            'result': t[7],
            'comment': t[8]
        })
    df = pd.DataFrame(data)
    df['direction'] = df['direction'].replace({'LONG': '🟢 LONG', 'SHORT': '🔴 SHORT'})
    df['result'] = df['result'].replace({'TAKE': '✅ Тейк', 'STOP': '❌ Стоп', 'BU': '⚖️ БУ'})
    
    fname = f"backtest_period_{period_id}.xlsx"
    with pd.ExcelWriter(fname, engine='openpyxl') as w:
        df.to_excel(w, sheet_name='Сделки', index=False)
        ws = w.sheets['Сделки']
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2b6cb0", end_color="2b6cb0", fill_type="solid")
        for col in range(1, len(df.columns)+1):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).fill = header_fill
        for col in df.columns:
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[get_column_letter(df.columns.get_loc(col)+1)].width = min(max_len, 30)
    
    await call.message.answer_document(document=FSInputFile(fname), caption=f"📊 Отчёт: {period[2]}")
    os.remove(fname)
    await call.answer()

# ---------- ДОБАВЛЕНИЕ ПЕРИОДА ----------
@dp.callback_query(F.data == "backtest_add_period")
async def bt_add_period(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(BacktestPeriodForm.name)
    await call.message.edit_text("📝 Введите название периода (например: Тест стратегии 1):", reply_markup=back_kb())
    await call.answer()

@dp.message(BacktestPeriodForm.name)
async def bt_period_name(msg: Message, state: FSMContext):
    await state.update_data(name=msg.text.strip())
    await state.set_state(BacktestPeriodForm.asset)
    await msg.answer("🪙 Введите актив для этого периода (например: BTCUSDT, EURUSD):", reply_markup=back_kb())

@dp.message(BacktestPeriodForm.asset)
async def bt_period_asset(msg: Message, state: FSMContext):
    await state.update_data(asset=msg.text.upper())
    await state.set_state(BacktestPeriodForm.period_start)
    await msg.answer("📅 Введите ДАТУ НАЧАЛА периода (ДД.ММ.ГГГГ):", reply_markup=back_kb())

@dp.message(BacktestPeriodForm.period_start)
async def bt_period_start(msg: Message, state: FSMContext):
    try:
        d = datetime.strptime(msg.text.strip(), "%d.%m.%Y").strftime("%Y-%m-%d")
        await state.update_data(period_start=d)
        await state.set_state(BacktestPeriodForm.period_end)
        await msg.answer("📅 Введите ДАТУ КОНЦА периода (ДД.ММ.ГГГГ):", reply_markup=back_kb())
    except ValueError:
        await msg.answer("❌ Ошибка! Введите дату в формате ДД.ММ.ГГГГ", reply_markup=back_kb())

@dp.message(BacktestPeriodForm.period_end)
async def bt_period_end(msg: Message, state: FSMContext):
    try:
        d = datetime.strptime(msg.text.strip(), "%d.%m.%Y").strftime("%Y-%m-%d")
        await state.update_data(period_end=d)
        await state.set_state(BacktestPeriodForm.initial_balance)
        await msg.answer("💰 Введите начальный баланс для этого периода (в $):", reply_markup=back_kb())
    except ValueError:
        await msg.answer("❌ Ошибка! Введите дату в формате ДД.ММ.ГГГГ", reply_markup=back_kb())

@dp.message(BacktestPeriodForm.initial_balance)
async def bt_initial_balance(msg: Message, state: FSMContext):
    try:
        balance = float(msg.text.replace(",", "."))
        data = await state.get_data()
        period_id = save_backtest_period(
            user_id=msg.from_user.id,
            name=data['name'],
            asset=data['asset'],
            initial_balance=balance,
            period_start=data['period_start'],
            period_end=data['period_end']
        )
        await state.clear()
        await msg.answer(f"✅ Период '{data['name']}' создан!\n\n🪙 Актив: {data['asset']}\n📅 {data['period_start']} — {data['period_end']}\n💰 Баланс: ${balance:.2f}\n\nТеперь вы можете добавлять сделки в этот период.", reply_markup=backtest_menu())
    except ValueError:
        await msg.answer("❌ Ошибка! Введите число.", reply_markup=back_kb())

# ---------- СТАТИСТИКА СПИСОК ПЕРИОДОВ ----------
@dp.callback_query(F.data == "backtest_stats_list")
async def bt_stats_list(call: CallbackQuery, state: FSMContext):
    periods = get_backtest_periods(call.from_user.id)
    if not periods:
        await call.answer("📭 Нет периодов", show_alert=True)
        return
    await state.update_data(stats_page=1)
    await show_periods_list(call, state, "stats")

@dp.callback_query(F.data.startswith("stats_page_"))
async def stats_periods_page(call: CallbackQuery, state: FSMContext):
    page = int(call.data.split("_")[2])
    await state.update_data(stats_page=page)
    await show_periods_list(call, state, "stats")

@dp.callback_query(F.data.startswith("stats_period_"))
async def bt_show_stats(call: CallbackQuery):
    period_id = int(call.data.split("_")[2])
    period = get_backtest_period_by_id(period_id, call.from_user.id)
    if not period:
        await call.answer("❌ Период не найден", show_alert=True)
        return
    trades = get_backtest_trades(period_id)
    text = get_backtest_stats_text(trades, period[4], period[2])
    await call.message.edit_text(text, reply_markup=backtest_menu())
    await call.answer()

# ---------- EXCEL СПИСОК ПЕРИОДОВ ----------
@dp.callback_query(F.data == "backtest_excel_list")
async def bt_excel_list(call: CallbackQuery, state: FSMContext):
    periods = get_backtest_periods(call.from_user.id)
    if not periods:
        await call.answer("📭 Нет периодов", show_alert=True)
        return
    await state.update_data(excel_page=1)
    await show_periods_list(call, state, "excel")

@dp.callback_query(F.data.startswith("excel_page_"))
async def excel_periods_page(call: CallbackQuery, state: FSMContext):
    page = int(call.data.split("_")[2])
    await state.update_data(excel_page=page)
    await show_periods_list(call, state, "excel")

@dp.callback_query(F.data.startswith("excel_period_"))
async def bt_export_excel(call: CallbackQuery):
    period_id = int(call.data.split("_")[2])
    period = get_backtest_period_by_id(period_id, call.from_user.id)
    if not period:
        await call.answer("❌ Период не найден", show_alert=True)
        return
    trades = get_backtest_trades(period_id)
    
    if not trades:
        await call.answer("📭 Нет сделок для экспорта", show_alert=True)
        return
    
    data = []
    for t in trades:
        data.append({
            'trade_date': t[1],
            'direction': t[2],
            'entry_price': t[3],
            'exit_price': t[4],
            'volume': t[5],
            'pnl': t[6],
            'result': t[7],
            'comment': t[8]
        })
    df = pd.DataFrame(data)
    df = df[['trade_date', 'direction', 'entry_price', 'exit_price', 'volume', 'pnl', 'result', 'comment']]
    df.columns = ['📅 Дата', '📈 Направление', '💰 Вход', '💰 Выход', '📊 Объём', '💵 P&L', '🎯 Исход', '📝 Комментарий']
    df['📈 Направление'] = df['📈 Направление'].replace({'LONG': '🟢 LONG', 'SHORT': '🔴 SHORT'})
    df['🎯 Исход'] = df['🎯 Исход'].replace({'TAKE': '✅ Тейк', 'STOP': '❌ Стоп', 'BU': '⚖️ БУ'})
    df = df.sort_values('📅 Дата', ascending=False)
    
    period_info = pd.DataFrame([{
        'period_name': period[2],
        'asset': period[3],
        'initial_balance': period[4],
        'period_start': period[5],
        'period_end': period[6]
    }])
    
    fname = f"backtest_period_{period_id}.xlsx"
    with pd.ExcelWriter(fname, engine='openpyxl') as w:
        period_info.to_excel(w, sheet_name='📊 Информация', index=False)
        df.to_excel(w, sheet_name='📋 Сделки', index=False)
        ws = w.sheets['📋 Сделки']
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2b6cb0", end_color="2b6cb0", fill_type="solid")
        for col in range(1, len(df.columns)+1):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).fill = header_fill
        for col in range(1, len(df.columns)+1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in range(1, len(df)+2):
                v = ws.cell(row=row, column=col).value
                if v:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len+2, 30)
        ws.freeze_panes = 'A2'
    
    await call.message.answer_document(document=FSInputFile(fname), caption=f"📊 Отчёт бэктеста: {period[2]}")
    os.remove(fname)
    await call.answer()

# ---------- ГРАФИК ПЕРИОДА ----------
@dp.callback_query(F.data.startswith("bt_chart_"))
async def bt_show_chart(call: CallbackQuery):
    period_id = int(call.data.split("_")[2])
    period = get_backtest_period_by_id(period_id, call.from_user.id)
    if not period:
        await call.answer("❌ Период не найден", show_alert=True)
        return
    trades = get_backtest_trades(period_id)
    if not trades:
        await call.answer("📭 Нет сделок для построения графика", show_alert=True)
        return
    
    chart_path = generate_backtest_equity_chart(trades, period[4], period_id, period[2])
    if chart_path:
        await call.message.answer_photo(photo=FSInputFile(chart_path), caption=f"📈 Кривая доходности: {period[2]}")
        os.remove(chart_path)
    await call.answer()

# ---------- ДОБАВЛЕНИЕ СДЕЛКИ В ПЕРИОД ----------
@dp.callback_query(F.data.startswith("bt_add_trade_"))
async def bt_add_trade(call: CallbackQuery, state: FSMContext):
    period_id = int(call.data.split("_")[3])
    await state.clear()
    await state.update_data(period_id=period_id)
    await state.set_state(BacktestTradeForm.trade_date)
    await call.message.edit_text("📅 Введите дату сделки (ДД.ММ.ГГГГ) или 'сегодня':", reply_markup=cancel_backtest_kb())
    await call.answer()

@dp.message(BacktestTradeForm.trade_date)
async def bt_trade_date(msg: Message, state: FSMContext):
    dstr = msg.text.strip().lower()
    if dstr in ["сегодня", "today"]:
        trade_date = datetime.now().strftime("%Y-%m-%d")
    else:
        try:
            trade_date = datetime.strptime(dstr, "%d.%m.%Y").strftime("%Y-%m-%d")
        except ValueError:
            await msg.answer("❌ Ошибка! Введите дату в формате ДД.ММ.ГГГГ", reply_markup=cancel_backtest_kb())
            return
    await state.update_data(trade_date=trade_date)
    await state.set_state(BacktestTradeForm.direction)
    await msg.answer("📈 Выберите направление:", reply_markup=backtest_direction_kb())

@dp.callback_query(F.data == "bt_dir_LONG")
async def bt_trade_direction_long(call: CallbackQuery, state: FSMContext):
    await state.update_data(direction="LONG")
    await state.set_state(BacktestTradeForm.entry_price)
    await call.message.edit_text("💰 Введите цену входа:", reply_markup=cancel_backtest_kb())
    await call.answer()

@dp.callback_query(F.data == "bt_dir_SHORT")
async def bt_trade_direction_short(call: CallbackQuery, state: FSMContext):
    await state.update_data(direction="SHORT")
    await state.set_state(BacktestTradeForm.entry_price)
    await call.message.edit_text("💰 Введите цену входа:", reply_markup=cancel_backtest_kb())
    await call.answer()

@dp.message(BacktestTradeForm.entry_price)
async def bt_trade_entry(msg: Message, state: FSMContext):
    try:
        val = parse_number(msg.text)
        if val <= 0:
            raise ValueError
        await state.update_data(entry_price=val)
        await state.set_state(BacktestTradeForm.exit_price)
        await msg.answer("💰 Введите цену выхода:", reply_markup=cancel_backtest_kb())
    except ValueError:
        await msg.answer("❌ Ошибка! Введите число (например: 56700.50 или 0.25)", reply_markup=cancel_backtest_kb())

@dp.message(BacktestTradeForm.exit_price)
async def bt_trade_exit(msg: Message, state: FSMContext):
    try:
        val = parse_number(msg.text)
        if val <= 0:
            raise ValueError
        await state.update_data(exit_price=val)
        await state.set_state(BacktestTradeForm.volume)
        await msg.answer("📊 Введите объём позиции (например: 0.25 или 1000):", reply_markup=cancel_backtest_kb())
    except ValueError:
        await msg.answer("❌ Ошибка! Введите число (например: 56700.50 или 0.25)", reply_markup=cancel_backtest_kb())

@dp.message(BacktestTradeForm.volume)
async def bt_trade_volume(msg: Message, state: FSMContext):
    try:
        vol = parse_number(msg.text)
        if vol <= 0:
            raise ValueError
        data = await state.get_data()
        direction = data['direction']
        entry = data['entry_price']
        exit_p = data['exit_price']
        pnl = (exit_p - entry) * vol if direction == "LONG" else (entry - exit_p) * vol
        await state.update_data(volume=vol, pnl=pnl)
        await state.set_state(BacktestTradeForm.result)
        await msg.answer("🎯 Как закрылась сделка?", reply_markup=backtest_result_kb())
    except ValueError:
        await msg.answer("❌ Ошибка! Введите число (например: 0.25 или 1000)", reply_markup=cancel_backtest_kb())

@dp.callback_query(F.data == "bt_res_TAKE")
async def bt_trade_result_take(call: CallbackQuery, state: FSMContext):
    await state.update_data(result="TAKE")
    await state.set_state(BacktestTradeForm.comment)
    await call.message.edit_text("📝 Введите комментарий (отправьте '-' чтобы пропустить):", reply_markup=cancel_backtest_kb())
    await call.answer()

@dp.callback_query(F.data == "bt_res_STOP")
async def bt_trade_result_stop(call: CallbackQuery, state: FSMContext):
    await state.update_data(result="STOP")
    await state.set_state(BacktestTradeForm.comment)
    await call.message.edit_text("📝 Введите комментарий (отправьте '-' чтобы пропустить):", reply_markup=cancel_backtest_kb())
    await call.answer()

@dp.callback_query(F.data == "bt_res_BU")
async def bt_trade_result_bu(call: CallbackQuery, state: FSMContext):
    await state.update_data(result="BU")
    await state.update_data(pnl=0)
    await state.set_state(BacktestTradeForm.comment)
    await call.message.edit_text("📝 Введите комментарий (отправьте '-' чтобы пропустить):", reply_markup=cancel_backtest_kb())
    await call.answer()

@dp.message(BacktestTradeForm.comment)
async def bt_trade_comment(msg: Message, state: FSMContext):
    com = msg.text.strip()
    await state.update_data(comment="" if com == "-" else com)
    await state.set_state(BacktestTradeForm.add_link)
    await msg.answer("🔗 Хотите добавить ссылку на график?", reply_markup=backtest_link_kb())

@dp.callback_query(F.data == "bt_add_link")
async def bt_add_link(call: CallbackQuery, state: FSMContext):
    await state.set_state(BacktestTradeForm.link_timeframe)
    await call.message.edit_text("📊 Выберите таймфрейм:", reply_markup=backtest_timeframe_kb())
    await call.answer()

@dp.callback_query(F.data == "bt_link_done")
async def bt_link_done(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    
    save_backtest_trade(
        period_id=data['period_id'],
        trade_date=data['trade_date'],
        direction=data['direction'],
        entry_price=data['entry_price'],
        exit_price=data['exit_price'],
        volume=data['volume'],
        pnl=data['pnl'],
        result=data['result'],
        comment=data['comment']
    )
    
    await state.clear()
    await call.message.edit_text("✅ Сделка добавлена в бэктест!", reply_markup=backtest_menu())
    await call.answer()

@dp.callback_query(F.data.startswith("bt_tf_"))
async def bt_select_timeframe(call: CallbackQuery, state: FSMContext):
    tf = call.data.split("_")[2]
    await state.update_data(link_timeframe=tf)
    await state.set_state(BacktestTradeForm.link_url)
    await call.message.edit_text(f"🔗 Введите ссылку для таймфрейма {tf}:", reply_markup=cancel_backtest_kb())
    await call.answer()

@dp.message(BacktestTradeForm.link_url)
async def bt_save_link(msg: Message, state: FSMContext):
    link = msg.text.strip()
    if not (link.startswith("http://") or link.startswith("https://")):
        await msg.answer("❌ Ошибка! Ссылка должна начинаться с http:// или https://", reply_markup=cancel_backtest_kb())
        return
    
    data = await state.get_data()
    period_id = data['period_id']
    tf = data['link_timeframe']
    save_backtest_link(period_id, tf, link)
    
    await state.set_state(BacktestTradeForm.add_link)
    await msg.answer("✅ Ссылка сохранена! Хотите добавить ещё?", reply_markup=backtest_link_kb())

# ---------- УДАЛЕНИЕ ПЕРИОДА ----------
@dp.callback_query(F.data.startswith("bt_clear_period_"))
async def bt_clear_period_confirm(call: CallbackQuery, state: FSMContext):
    period_id = int(call.data.split("_")[3])
    await state.update_data(delete_period_id=period_id)
    await call.message.edit_text(f"⚠️ Удалить период и все его сделки?", reply_markup=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да", callback_data="bt_clear_period_yes"),
         InlineKeyboardButton(text="❌ Нет", callback_data="backtest_list_periods")]
    ]))
    await call.answer()

@dp.callback_query(F.data == "bt_clear_period_yes")
async def bt_clear_period_execute(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    period_id = data.get('delete_period_id')
    if period_id:
        delete_backtest_period(period_id, call.from_user.id)
    await state.clear()
    await call.message.edit_text("🗑 Период удалён!", reply_markup=backtest_menu())
    await call.answer()

# ==================================================
# БЛОК 15: КОМАНДЫ
# ==================================================

@dp.message(Command("new"))
async def cmd_new(msg: Message, state: FSMContext):
    await state.clear()
    await state.set_state(TradeForm.asset)
    await msg.answer("Введите тикер (BTC, ETH, TON, AAPL):", reply_markup=cancel_real_kb())

@dp.message(Command("stats"))
async def cmd_stats(msg: Message):
    df = get_trades_filtered(msg.from_user.id)
    text = get_stats_text(df)
    chart = generate_equity_chart(df, msg.from_user.id)
    if chart:
        await msg.answer_photo(photo=FSInputFile(chart), caption=text)
        os.remove(chart)
    else:
        await msg.answer(text)

@dp.message(Command("day"))
async def cmd_day(msg: Message):
    df = get_trades_filtered(msg.from_user.id, date_filter="day")
    text = get_stats_text_short(df, "Статистика за сегодня")
    await msg.answer(text)

@dp.message(Command("week"))
async def cmd_week(msg: Message):
    df = get_trades_filtered(msg.from_user.id, date_filter="week")
    text = get_stats_text_short(df, "Статистика за неделю")
    await msg.answer(text)

@dp.message(Command("month"))
async def cmd_month(msg: Message):
    df = get_trades_filtered(msg.from_user.id, date_filter="month")
    text = get_stats_text_short(df, "Статистика за месяц")
    await msg.answer(text)

@dp.message(Command("clear"))
async def cmd_clear(msg: Message):
    clear_trades(msg.from_user.id)
    await msg.answer("Журнал реальной торговли очищен!")

@dp.message(Command("change_language"))
async def cmd_change_language(msg: Message):
    await msg.answer("Выберите язык / Choose language:", reply_markup=lang_kb())

# ==================================================
# БЛОК 16: ЗАПУСК
# ==================================================
async def set_commands():
    await bot.set_my_commands([
        BotCommand(command="start", description="Главное меню"),
        BotCommand(command="new", description="➕ Новая сделка"),
        BotCommand(command="stats", description="📊 Вся статистика"),
        BotCommand(command="day", description="📆 За сегодня"),
        BotCommand(command="week", description="📅 За неделю"),
        BotCommand(command="month", description="📊 За месяц"),
        BotCommand(command="clear", description="🗑 Очистить журнал"),
        BotCommand(command="change_language", description="🌐 Сменить язык"),
    ])

async def main():
    # Удаляем вебхук на всякий случай
    try:
        await bot.delete_webhook()
        print("✅ Вебхук удалён")
    except:
        pass
    
    init_dbs()
    await set_commands()
    print("✅ Бот запущен и работает через polling")
    
    # Только polling, без веб-сервера
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
